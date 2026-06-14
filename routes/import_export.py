import io, csv, secrets, string
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime

from models import get_db, User, Enrollment, ExamSubmission, Exam, Homework, HomeworkSubmission
from auth import require_admin, hash_password

router = APIRouter(prefix="/api/import-export", tags=["import-export"])


# ── Génération de mot de passe ────────────────────
def _gen_password(length=10) -> str:
    chars = string.ascii_letters.replace('l','').replace('O','') + string.digits.replace('0','').replace('1','')
    return ''.join(secrets.choice(chars) for _ in range(length))


# ══════════════════════════════════════════════════
#  IMPORT MASSIF ÉTUDIANTS
# ══════════════════════════════════════════════════

@router.post("/students/import")
async def import_students(
    file:       UploadFile = File(...),
    course_id:  int = None,
    db:         Session = Depends(get_db),
    me:         User    = Depends(require_admin),
):
    """
    Importe des étudiants depuis un fichier CSV ou Excel.
    Colonnes attendues : name, email (password optionnel)
    Si course_id fourni, inscrit automatiquement au cours.
    """
    ext = file.filename.lower().split(".")[-1] if file.filename else ""
    contents = await file.read()

    rows = []

    if ext == "csv":
        text = contents.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

    elif ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows())]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else "" for v in row])))
        except ImportError:
            raise HTTPException(500, "openpyxl non installé — pip install openpyxl")
    else:
        raise HTTPException(400, "Format non supporté. Utilisez CSV ou XLSX.")

    created = []
    updated = []
    errors  = []
    credentials = []

    for i, row in enumerate(rows, start=2):
        name  = (row.get("name") or row.get("nom") or row.get("prénom nom") or "").strip()
        email = (row.get("email") or row.get("mail") or "").strip().lower()

        if not name or not email:
            errors.append({"ligne": i, "raison": "Nom ou email manquant"})
            continue

        if "@" not in email:
            errors.append({"ligne": i, "raison": f"Email invalide : {email}"})
            continue

        existing = db.query(User).filter_by(email=email).first()
        if existing:
            if course_id:
                enr = db.query(Enrollment).filter_by(student_id=existing.id, course_id=course_id).first()
                if not enr:
                    db.add(Enrollment(student_id=existing.id, course_id=course_id))
            updated.append(email)
            continue

        pwd = (row.get("password") or row.get("mot_de_passe") or "").strip() or _gen_password()
        user = User(
            name=name, email=email,
            hashed_pwd=hash_password(pwd),
            role="student", is_active=True,
        )
        db.add(user); db.flush()

        if course_id:
            db.add(Enrollment(student_id=user.id, course_id=course_id))

        credentials.append({"name": name, "email": email, "password": pwd})
        created.append(email)

    db.commit()

    return {
        "created":     len(created),
        "updated":     len(updated),
        "errors":      errors,
        "credentials": credentials,  # mots de passe générés — à communiquer aux étudiants
    }


@router.get("/students/template")
def download_template(me: User = Depends(require_admin)):
    """Télécharge un template CSV vide pour l'import d'étudiants."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["name", "email", "password"])
    writer.writerow(["Ahmadou Bello", "ahmadou@email.cm", ""])
    writer.writerow(["Fatouma Ali",   "fatouma@email.cm", "motdepasse123"])
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_etudiants.csv"},
    )


# ══════════════════════════════════════════════════
#  EXPORT RÉSULTATS EXAMENS
# ══════════════════════════════════════════════════

@router.get("/exams/{exam_id}/export")
def export_exam_results(
    exam_id: int,
    db:      Session = Depends(get_db),
    me:      User    = Depends(require_admin),
):
    """Exporte les résultats d'un examen en Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl non installé — pip install openpyxl")

    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(404, "Examen introuvable")

    subs = db.query(ExamSubmission).filter_by(exam_id=exam_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats"

    # Style en-tête
    header_fill = PatternFill("solid", fgColor="1E4DB7")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Nom étudiant", "Email", "Score", "Score max", "Pourcentage", "Mention", "Date soumission", "Statut", "Violations"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    # Données
    for row_idx, sub in enumerate(subs, start=2):
        student = sub.student
        pct = round(sub.score / sub.max_score * 100, 1) if sub.score is not None and sub.max_score else 0

        if pct >= 80:   mention = "Très bien"
        elif pct >= 70: mention = "Bien"
        elif pct >= 60: mention = "Assez bien"
        elif pct >= 50: mention = "Passable"
        else:           mention = "Insuffisant"

        passing = exam.passing_score or 50
        status  = "Reçu" if pct >= passing else "Ajourné"

        ws.append([
            student.name if student else "Inconnu",
            student.email if student else "",
            sub.score or 0,
            sub.max_score or 0,
            f"{pct}%",
            mention,
            sub.submitted_at.strftime("%d/%m/%Y %H:%M") if sub.submitted_at else "",
            status,
            sub.violations or 0,
        ])

        # Colorier selon mention
        color = "D1FAE5" if pct >= 50 else "FEE2E2"
        fill  = PatternFill("solid", fgColor=color)
        for col in range(1, len(headers)+1):
            ws.cell(row=row_idx, column=col).fill = fill

    # Feuille récapitulatif
    ws2 = wb.create_sheet("Récapitulatif")
    scores = [s.score for s in subs if s.score is not None]
    recus  = sum(1 for s in subs if s.score and s.max_score and (s.score/s.max_score*100) >= (exam.passing_score or 50))

    ws2.append(["Statistiques", ""])
    ws2.append(["Examen",          exam.title])
    ws2.append(["Total soumissions", len(subs)])
    ws2.append(["Reçus",           recus])
    ws2.append(["Ajournés",        len(subs) - recus])
    ws2.append(["Moyenne (%)",     f"{round(sum(s/m*100 for s,m in [(s.score,s.max_score) for s in subs if s.score and s.max_score])/len(subs),1)}%" if subs else "N/A"])
    ws2.append(["Score min",       min(scores) if scores else "N/A"])
    ws2.append(["Score max",       max(scores) if scores else "N/A"])
    ws2.append(["Généré le",       datetime.now().strftime("%d/%m/%Y %H:%M")])

    for row in ws2.iter_rows():
        row[0].font = Font(bold=True)

    # Sauvegarder en mémoire
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"resultats_{exam.title.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════
#  EXPORT RÉSULTATS DEVOIRS
# ══════════════════════════════════════════════════

@router.get("/homeworks/{hw_id}/export")
def export_homework_results(
    hw_id: int,
    db:    Session = Depends(get_db),
    me:    User    = Depends(require_admin),
):
    """Exporte les résultats d'un devoir en Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl non installé")

    hw = db.query(Homework).filter(Homework.id == hw_id).first()
    if not hw:
        raise HTTPException(404, "Devoir introuvable")

    subs = db.query(HomeworkSubmission).filter_by(homework_id=hw_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats devoirs"

    header_fill = PatternFill("solid", fgColor="10B981")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Nom étudiant", "Email", "Note", "Note max", "Rendu en retard", "Commentaire étudiant", "Feedback enseignant", "Date soumission", "Statut"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    for row_idx, sub in enumerate(subs, start=2):
        student = sub.student
        ws.append([
            student.name  if student else "Inconnu",
            student.email if student else "",
            sub.score if sub.score is not None else "Non noté",
            hw.max_score,
            "Oui" if sub.late else "Non",
            sub.comment or "",
            sub.feedback or "",
            sub.submitted_at.strftime("%d/%m/%Y %H:%M") if sub.submitted_at else "",
            "Corrigé" if sub.graded else "En attente",
        ])
        color = "D1FAE5" if sub.graded else "FEF9C3"
        fill  = PatternFill("solid", fgColor=color)
        for col in range(1, len(headers)+1):
            ws.cell(row=row_idx, column=col).fill = fill

    output = io.BytesIO()
    wb.save(output); output.seek(0)

    filename = f"devoir_{hw.title.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════
#  HISTORIQUE CONNEXIONS
# ══════════════════════════════════════════════════

@router.get("/login-history")
def get_login_history(
    user_id: int = None,
    limit:   int = 100,
    db:      Session = Depends(get_db),
    me:      User    = Depends(require_admin),
):
    """Export historique des connexions en Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl non installé")

    from models import LoginHistory
    q = db.query(LoginHistory)
    if user_id:
        q = q.filter_by(user_id=user_id)
    logs = q.order_by(LoginHistory.created_at.desc()).limit(limit).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Connexions"

    headers = ["Utilisateur", "Email", "IP", "Navigateur", "Succès", "Date"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = 20

    for log in logs:
        user = log.user
        ws.append([
            user.name  if user else "Inconnu",
            user.email if user else "",
            log.ip_address or "",
            (log.user_agent or "")[:80],
            "Oui" if log.success else "Non",
            log.created_at.strftime("%d/%m/%Y %H:%M") if log.created_at else "",
        ])

    output = io.BytesIO()
    wb.save(output); output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=historique_connexions_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )
